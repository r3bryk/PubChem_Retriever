# Import libraries, packages, and functions
import pandas as pd
import pubchempy as pcp
import tkinter as tk
from tkinter import Tk, filedialog, simpledialog
import os
import re
import time
import requests
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.worksheet.dimensions import ColumnDimension

# Toggles
def prompt_user(question):
    while True:
        response = input(f"{question} \n1 = Yes \n0 = No \nType your answer and press Enter: ").strip()
        if response == '1':
            print("-"*150)
            return True
        elif response == '0':
            print("-"*150)
            return False
        else:
            print("Invalid input. Please enter 1 for Yes or 2 for No.")

# Interactive prompts for data retrieval options
RETRIEVE_INCHIKEY = prompt_user("Would you like to retrieve InChIKeys from PubChem?")
RETRIEVE_CAS = prompt_user("Would you like to retrieve CAS# from PubChem?")
RETRIEVE_SMILES = prompt_user("Would you like to retrieve SMILES from PubChem?")
RETRIEVE_DTXSID = prompt_user("Would you like to retrieve DSSTox SID from PubChem?")
RETRIEVE_USES = prompt_user("Would you like to retrieve Uses & Use Classification from PubChem?")
RETRIEVE_CLASSYFIRE = prompt_user("Would you like to retrieve chemical class data from ClassyFire Batch?")


def get_pubchem_info(identifier, input_type='name', retrieve_inchikey=True, retrieve_cas=True, retrieve_smiles=True):
    """
    Retrieve InChIKey, CAS, and SMILES using PubChem property API based on CID,
    which is obtained from either name or InChIKey.
    """
    try:
        # First get PubChem CID
        if input_type == 'name':
            url_cid = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{identifier}/cids/JSON"
        else:
            url_cid = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/inchikey/{identifier}/cids/JSON"
        res = requests.get(url_cid, timeout=10)
        res.raise_for_status()
        cids = res.json().get('IdentifierList', {}).get('CID', [])
        if not cids:
            return None, None, None
        cid = cids[0]

        # Fetch properties including SMILES and InChIKey
        prop_url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/property/InChIKey,SMILES/JSON"
        prop_res = requests.get(prop_url, timeout=10)
        prop_res.raise_for_status()
        props = prop_res.json().get('PropertyTable', {}).get('Properties', [{}])[0]

        inchikey = props.get('InChIKey') if retrieve_inchikey else None
        smiles = props.get('SMILES') if retrieve_smiles else None

        # CAS number via PUG-View
        cas = get_cas(cid) if retrieve_cas else None

        return inchikey, cas, smiles

    except Exception as e:
        print(f"Error fetching for '{identifier}' ({input_type}): {e}")
        return None, None, None

# A function to get CAS number
def get_cas(cid):
    """
    Retrieve CAS number using PubChem PUG-View and recursive heading search.
    """
    try:
        url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{cid}/JSON"
        res = requests.get(url, timeout=10)
        res.raise_for_status()
        data = res.json()
        sections = data.get('Record', {}).get('Section', [])
        cas_list = extract_heading_info(sections, 'CAS')
        for s in cas_list:
            match = re.search(r'\d{2,7}-\d{2}-\d', s)
            if match:
                return match.group(0)
    except:
        pass
    return None

# A function to get PubChem CID
def get_pubchem_cid(identifier, input_type='name'):
    try:
        if input_type == 'name':
            url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{identifier}/cids/JSON"
        elif input_type == 'inchikey':
            url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/inchikey/{identifier}/cids/JSON"
        else:
            return None

        res = requests.get(url)
        if res.status_code == 200:
            return res.json()['IdentifierList']['CID'][0]
    except:
        pass
    return None

# A function to get PubChem heading information
def extract_heading_info(sections, target_heading):
    results = []
    for section in sections:
        if section.get("TOCHeading") == target_heading:
            for info in section.get("Information", []):
                for val in info.get("Value", {}).get("StringWithMarkup", []):
                    s = val.get("String", "")
                    if s:
                        results.append(s)
        # Recurse into subsections
        if "Section" in section:
            results += extract_heading_info(section["Section"], target_heading)
    return results

# A function to find Uses and Use Classification sections
def find_use_sections(root_sections):
    uses_list, class_list = [], []
    for section in root_sections:
        if section.get("TOCHeading") == "Use and Manufacturing":
            # Under that, look for "Uses" and "Use Classification"
            subsec = section.get("Section", [])
            uses_list += extract_heading_info(subsec, "Uses")
            class_list += extract_heading_info(subsec, "Use Classification")
        # Recurse to catch nested
        if "Section" in section:
            u, c = find_use_sections(section["Section"])
            uses_list += u
            class_list += c
    return uses_list, class_list

# A function to get Uses and Use Classification
def get_pubchem_uses(cid):
    uses, classifications = [], []
    try:
        url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{cid}/JSON"
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        data = r.json()
        root = data.get("Record", {}).get("Section", [])
        uses, classifications = find_use_sections(root)
    except Exception as e:
        print(f"Error retrieving uses/classification for CID {cid}: {e}")
    return "; ".join(uses), "; ".join(classifications)

# A function to find DTXSID section
def extract_heading_dtxsid(sections, target_heading):
    results = []
    for section in sections:
        heading = section.get("TOCHeading", "")
        if heading == target_heading:
            for info in section.get("Information", []):
                val_obj = info.get("Value", {})
                if isinstance(val_obj, dict):
                    string_markup = val_obj.get("StringWithMarkup", [])
                    for val in string_markup:
                        string = val.get("String", "")
                        if string:
                            results.append(string)
        if "Section" in section:
            results += extract_heading_dtxsid(section["Section"], target_heading)
    return results

# A helper function to get DTXSID
def get_dtxsid(cid):
    if cid is None:
        return None
    try:
        url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{cid}/JSON"
        res = requests.get(url, timeout=10)
        res.raise_for_status()
        data = res.json()
        root_sections = data.get("Record", {}).get("Section", [])

        # Extract DSSTox Substance ID from its actual heading
        dtxsid_list = extract_heading_dtxsid(root_sections, "DSSTox Substance ID")
        for sid in dtxsid_list:
            if sid.startswith("DTXSID"):
                return sid
        return None
    except Exception:
        return None

# A function to get DTXSID
def get_pubchem_dtxsid(identifier, input_type='name'):
    cid = get_pubchem_cid(identifier, input_type)
    return get_dtxsid(cid)

# A function to get chemical class data from ClassyFire Batch
def retrieve_classyfire_classification(df):
    inchikey_col = (
        'InChIKey_Consensus' if 'InChIKey_Consensus' in df.columns else
        'InChIKey_PubChem' if 'InChIKey_PubChem' in df.columns else
        'InChIKey' if 'InChIKey' in df.columns else
        None
    )
    if inchikey_col is None:
        print("No InChIKey column found. Skipping ClassyFire retrieval.\n" + "!"*150)
        return df

    # Re-sorting table by 'InChIKey' column in descending fashion, N/As are moved to the bottom of the table
    # This step is needed for further classificetion script to skip N/As
    df = df.sort_values(by = inchikey_col, ascending = False, na_position = 'last')
    df = df.reset_index(drop = True)
    #QC printing
    print('Sorted df by InChIKey:\n', df)
    print(150*'-')

    # Create a list of InChIKeys
    inchikeys = df['InChIKey']

    print(f"ClassyFire classification started for {len(inchikeys)} InChIKeys...")

    # Launch browser
    driver = webdriver.Chrome()
    driver.get("https://cfb.fiehnlab.ucdavis.edu/")

    try:
        # Wait for the search text area and the ClassyFy button; the time can be adjusted based on page loading time
        search_textarea = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, 'inchikeys')))
        classify_button = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[type='submit']")))
        
        # Insert InChIKeys from the list
        search_textarea.clear() #clearing the input field
        # for key in df['InChIKey'].dropna().tolist(): #.dropna().tolist() is used to skip N/As, otherwise, the script crashes
        for key in inchikeys.dropna().tolist(): #.dropna().tolist() is used to skip N/As, otherwise, the script crashes    
            search_textarea.send_keys(key + "\n")

        # Click the ClassyFy button after enter all InChIKeys
        classify_button.click()

        # Wait for the results to load; the time can be adjusted based on page loading time,
        # 22000 sec are used to be sure the processing will be done before the script exits the webpage
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, 'results')))
        table = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, 'table')))
        reset_button = WebDriverWait(driver, 22000).until(EC.element_to_be_clickable((By.CLASS_NAME, 'btn-warning')))
        
        # Extract data from the result table
        table_rows = table.find_elements(By.TAG_NAME, "tr") # Find all rows
        data = [] # Create an empty array to store the data
        for row in table_rows:
            cols = row.find_elements(By.TAG_NAME, "td")
            cols = [col.text for col in cols]
            data.append(cols)
        print("-" * 150)

        # Create a list of column names; max. 11 columns are returned by ClassyFire
        column_names = ['InChIKey_ClassyFire', 'Status', 'Kingdom', 'Superclass', 'Class_ClassyFire', 'Subclass', 'Parent Level 1', 'Parent Level 2', 'Parent Level 3', 'Parent Level 4', 'Parent Level 5']
        # Adjust data to align with the specified column names
        filled_data = [row + [None] * (len(column_names) - len(row)) for row in data]
        # Convert data to df & add specified column names
        df_classyfy = pd.DataFrame(filled_data, columns = column_names)
        # Drop rows where all values are None, 'Status' and 'Kingdom' columns, and reset indices
        df_classyfy.dropna(how = 'all', inplace = True)
        
        #QC printing
        print('ClassiFyed table:\n', df_classyfy)
        print("-" * 150)

    finally:
        driver.quit()

    df = df.merge(df_classyfy, how='left', left_on=inchikey_col, right_on='InChIKey_ClassyFire')
    print("ClassyFire chemical classification added.\n" + "-"*150)
    df = df.drop(['InChIKey_ClassyFire', 'Status', 'Kingdom'], axis = 1)
    df.reset_index(drop = True, inplace = True)
    return df

# Main function
def main():
    print("-"*150)
    start_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print("Processing start time: ", start_time)
    print("-"*150)

    try:
        root = Tk()
        root.withdraw()
        root.update()
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("Text files", "*.txt")]
        )
        root.destroy()
    except Exception as e:
        print(f"File dialog failed ({e}). Please enter file path manually:")
        file_path = input("Path to input file: ").strip()

    if not file_path or not os.path.exists(file_path):
        print("No valid file selected or path does not exist.\n" + "!" * 150)
        return

    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.xlsx':
        df = pd.read_excel(file_path, engine='openpyxl')
    elif ext in ['.csv', '.txt']:
        df = pd.read_csv(file_path, sep=None, engine='python')
    else:
        print("Unsupported file format.\n" + "!"*150)
        return

    has_name = 'Name' in df.columns
    has_inchikey = 'InChIKey' in df.columns

    if not has_name and not has_inchikey:
        print("Neither 'Name' nor 'InChIKey' column found. Exiting.\n" + "!"*150)
        return

    # Decide input mode
    if has_name and has_inchikey:
        print("Choose input column for compound search:")
        print("1 = Name")
        print("2 = InChIKey")
        while True:
            choice = input("Type 1 or 2 and press Enter: ").strip()
            if choice in ['1', '2']:
                input_mode = 'name' if choice == '1' else 'inchikey'
                break
            else:
                print("Invalid input. Please enter 1 or 2.")
        print("-" * 150)
    else:
        input_mode = 'name' if has_name else 'inchikey'

    # Prepare output columns
    inchikeys, cass, smiles_list = [], [], []
    uses_list, use_class_list = [], []
    dtxsids = []

    # Input series
    primary_series = df['Name'] if input_mode == 'name' else df['InChIKey']
    fallback_series = df['InChIKey'] if input_mode == 'name' and has_inchikey else df['Name'] if input_mode == 'inchikey' and has_name else None
    fallback_mode = 'inchikey' if input_mode == 'name' else 'name'

    for idx, (i, primary_val) in enumerate(primary_series.items(), start=1):
        inchikey = cas = smiles = uses = use_class = dtxsid = None
        # Skip if it's an unknown feature based on Name column
        if 'Name' in df.columns and isinstance(df.at[i, 'Name'], str) and 'Feature' in df.at[i, 'Name']:
            print(f"{idx}: {primary_val}  -->  Skipped (unknown feature)")
        elif pd.isna(primary_val) or not isinstance(primary_val, str):
            print(f"{idx}: {primary_val}  -->  Skipped (invalid primary input)")
        else:
            inchikey, cas, smiles = get_pubchem_info(primary_val, input_type=input_mode,
                                                    retrieve_inchikey=RETRIEVE_INCHIKEY,
                                                    retrieve_cas=RETRIEVE_CAS,
                                                    retrieve_smiles=RETRIEVE_SMILES)
            if RETRIEVE_USES:
                cid = get_pubchem_cid(primary_val, input_type=input_mode)
                if cid:
                    uses, use_class = get_pubchem_uses(cid)
            if RETRIEVE_DTXSID:
                dtxsid = get_pubchem_dtxsid(primary_val, input_type=input_mode)

        # If nothing was retrieved and fallback is available
        if fallback_series is not None and (not any([inchikey, cas, smiles, uses, use_class, dtxsid])):
            try:
                fallback_val = fallback_series.loc[i]
            except KeyError:
                fallback_val = None

            if pd.notna(fallback_val) and isinstance(fallback_val, str):
                print(f"Name for ID {idx} not found in PubChem. Using InChIKey instead: {fallback_val}")
                inchikey, cas, smiles = get_pubchem_info(fallback_val, input_type=fallback_mode,
                                                        retrieve_inchikey=RETRIEVE_INCHIKEY,
                                                        retrieve_cas=RETRIEVE_CAS,
                                                        retrieve_smiles=RETRIEVE_SMILES)
                if RETRIEVE_USES:
                    cid = get_pubchem_cid(fallback_val, input_type=fallback_mode)
                    if cid:
                        uses, use_class = get_pubchem_uses(cid)
                if RETRIEVE_DTXSID:
                    dtxsid = get_pubchem_dtxsid(fallback_val, input_type=fallback_mode)

        # Append results
        inchikeys.append(inchikey)
        cass.append(cas)
        smiles_list.append(smiles)
        uses_list.append(uses if RETRIEVE_USES else None)
        use_class_list.append(use_class if RETRIEVE_USES else None)
        dtxsids.append(dtxsid if RETRIEVE_DTXSID else None)

        result = [f"{idx}: {primary_val}"]
        if RETRIEVE_INCHIKEY: result.append(f"InChIKey: {inchikey or 'None'}")
        if RETRIEVE_CAS: result.append(f"CAS#: {cas or 'None'}")
        if RETRIEVE_SMILES: result.append(f"SMILES: {smiles or 'None'}")
        if RETRIEVE_USES:
            result.append(f"Uses: {'Found' if uses else 'None'}")
            result.append(f"Use Classification: {'Found' if use_class else 'None'}")
        if RETRIEVE_DTXSID:
            result.append(f"DTXSID: {dtxsid or 'None'}")

        print("; ".join(result))
        print("-" * 200)
        time.sleep(0.1)

    # Add results to DataFrame
    if RETRIEVE_INCHIKEY:
        df['InChIKey_PubChem'] = inchikeys
    if RETRIEVE_CAS:
        df['CAS_PubChem'] = cass
    if RETRIEVE_SMILES:
        df['SMILES_PubChem'] = smiles_list
    if RETRIEVE_USES:
        df['Uses'] = uses_list
        df['Use Classification'] = use_class_list
    if RETRIEVE_DTXSID:
        df['DTXSID_PubChem'] = dtxsids

    # Consensus InChIKey generation
    has_inchikey = 'InChIKey' in df.columns
    has_inchikey_pubchem = 'InChIKey_PubChem' in df.columns

    if not has_inchikey and not has_inchikey_pubchem:
        print("Neither 'InChIKey' nor 'InChIKey_PubChem' found. Skipping consensus InChIKey creation.")
        print("!"*150)
    else:
        # Determine insertion index
        insert_after_col = 'InChIKey' if has_inchikey else 'InChIKey_PubChem'
        insert_index = df.columns.get_loc(insert_after_col) + 1

        # Insert empty column
        df.insert(insert_index, 'InChIKey_Consensus', pd.NA)
        print(f"'InChIKey_Consensus' column inserted after '{insert_after_col}'.")

        # Fill from 'InChIKey_PubChem' first
        if has_inchikey_pubchem:
            df['InChIKey_Consensus'] = df['InChIKey_PubChem']
            print("Values from 'InChIKey_PubChem' copied to 'InChIKey_Consensus'.")

        # Fill NaNs with values from 'InChIKey'
        if has_inchikey:
            mask = df['InChIKey_Consensus'].isna()
            df.loc[mask, 'InChIKey_Consensus'] = df.loc[mask, 'InChIKey']
            print("NaN values in 'InChIKey_Consensus' filled using 'InChIKey'.")
            print("-"*150)
    
    # Add ClassyFire classification if requested
    if RETRIEVE_CLASSYFIRE:
        df = retrieve_classyfire_classification(df)

    # Reorder classification and PubChem columns
    columns_to_move = [
        'InChIKey_PubChem', 'CAS_PubChem', 'SMILES_PubChem', 'DTXSID_PubChem', 'Uses', 'Use Classification',
        'Superclass', 'Class_ClassyFire', 'Subclass',
        'Parent Level 1', 'Parent Level 2', 'Parent Level 3', 'Parent Level 4', 'Parent Level 5'
    ]
    
    # Filter to include only columns that actually exist in the DataFrame
    columns_to_move = [col for col in columns_to_move if col in df.columns]

    # Pop all columns first to avoid index issues
    popped_cols = {col: df.pop(col) for col in columns_to_move}

    # Append them at the end
    for col in columns_to_move:
        df[col] = popped_cols[col]
        
    # Format output Excel file
    def apply_excel_formatting(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Find header row and build column index
        header = [cell.value for cell in ws[1]]
        col_idx = {col: idx + 1 for idx, col in enumerate(header)}
        
        # Prevent Excel from converting CAS# to date/time
        if "CAS" in col_idx:
            col_letter = openpyxl.utils.get_column_letter(col_idx["CAS"])
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                cell.number_format = "@"
        if "CASRN" in col_idx:
            col_letter = openpyxl.utils.get_column_letter(col_idx["CASRN"])
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                cell.number_format = "@"
        if "CAS#" in col_idx:
            col_letter = openpyxl.utils.get_column_letter(col_idx["CAS#"])
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                cell.number_format = "@"
        if "CAS_PubChem" in col_idx:
            col_letter = openpyxl.utils.get_column_letter(col_idx["CAS_PubChem"])
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                cell.number_format = "@"

        # Headers alignment and row height
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[1].height = 60

        # Freeze top row
        ws.freeze_panes = "A2"

        # Set zoom level to 80%
        ws.sheet_view.zoomScale = 80

        # Set column widths safely
        for cols, width in [
            (["Name"], 52),
        ]:
            for col in cols:
                if col in col_idx:
                    col_letter = openpyxl.utils.get_column_letter(col_idx[col])
                    ws.column_dimensions[col_letter].width = width

        # Auto-adjust selected columns (approximate by content length)
        auto_cols = ["Name", "Formula", "CAS#", "CAS", "CAS_PubChem", "CASRN", "Class_FUse", "Class_NORMAN", "Class", "DTXSID_PubChem", "DTXSID"]
        for col in auto_cols:
            if col in col_idx:
                col_letter = openpyxl.utils.get_column_letter(col_idx[col])
                max_length = 0
                for row in ws.iter_rows(min_row=2, min_col=col_idx[col], max_col=col_idx[col]):
                    for cell in row:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                adjusted_width = min(max_length + 2, 100)
                ws.column_dimensions[col_letter].width = adjusted_width

        # Save workbook
        wb.save(file_path)
        print(f"Formatting applied and saved to: {file_path}")
        print(150 * "-")
        
    # Save results
    if file_path:
        print(f"Writing data to the output file...")
        base, ext = os.path.splitext(file_path)
        processed_path = f"{base}_PubChem.xlsx"
        df.to_excel(processed_path, index=False, engine='openpyxl')

        # Apply formatting
        apply_excel_formatting(processed_path)

        print(f"Processed data saved to: {processed_path}")
        print("-"*150)
        print("Processing start time: ", start_time)
        print("Processing end time: ", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        print("-"*150)
        print(150*"-")
    else:
        print(150*"!")
        print("No file selected.")
        print(150*"!")

if __name__ == "__main__":
    main()